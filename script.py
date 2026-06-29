import pandas as pd
import streamlit as st
import camelot
import tempfile
import os
import uuid

from io import BytesIO
from datetime import datetime

from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from utils import load_history, save_history, HISTORY_FOLDER


# =====================================================
# HELPERS
# =====================================================
def has_common_word(name1, name2):
    words1 = set(str(name1).upper().split())
    words2 = set(str(name2).upper().split())
    return len(words1 & words2) > 0

def safe_find_col(df, candidates):
    for c in df.columns:
        if any(c.lower().strip() == cand.lower() for cand in candidates):
            return c
    return None


def get_col_keyword(df, keywords):
    for col in df.columns:
        for k in keywords:
            if k.lower() in col.lower():
                return col
    return None


def ensure_numeric(series):
    return pd.to_numeric(series, errors="coerce").fillna(0)


def extract_levy(df, type_col, label, keyword):
    row = df[df[type_col] == label]
    if row.empty:
        return 0.0
    for col in df.columns:
        if keyword.lower() in str(col).lower():
            val = pd.to_numeric(row.iloc[0][col], errors="coerce")
            return 0.0 if pd.isna(val) else float(val)
    return 0.0


def extract_levies(df, type_col, label, keywords):
    if df is None or type_col not in df.columns:
        return 0.0
    row = df[df[type_col] == label]
    if row.empty:
        return 0.0
    total = 0.0
    for col in df.columns:
        if any(k.lower() in str(col).lower() for k in keywords):
            val = pd.to_numeric(row.iloc[0][col], errors="coerce")
            if not pd.isna(val):
                total += float(val)
    return round(total, 2)


# =====================================================
# MATCHING HELPER FUNCTIONS
# =====================================================
def norm_sec(val):
    """Normalise security ticker: 'DLTA.ZW' → 'DLTA'"""
    return str(val).upper().split(".")[0].strip()


def norm_date(val):
    """Normalise dates from any common format to YYYY-MM-DD string."""
    val = str(val).strip()
    # Handle Excel serial date floats that come through as strings like "46000.0"
    try:
        serial = float(val)
        if 30000 < serial < 60000:
            from datetime import date
            base = date(1899, 12, 30)
            import datetime as dt
            return (base + dt.timedelta(days=int(serial))).strftime("%Y-%m-%d")
    except (ValueError, TypeError):
        pass
    for fmt in (
        "%d-%b-%Y",   # 14-May-2026
        "%Y/%m/%d",   # 2026/05/14
        "%Y-%m-%d",   # 2026-05-14
        "%d/%m/%Y",   # 14/05/2026
        "%m/%d/%Y",   # 05/14/2026
        "%d-%m-%Y",   # 14-05-2026
    ):
        try:
            return datetime.strptime(val, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return val  # fallback: keep as-is


def norm_name(val):
    """Strip non-alpha chars for fuzzy name comparison."""
    return str(val).strip().upper().replace(r"[^A-Z]", "")


# =====================================================
# CDC RECEIPTING UI
# =====================================================
def cdc_receipting_ui():

    # ✅ Header with refresh
    col1, col2 = st.columns([9, 1])

    with col2:
        if st.button("🔄", key="cdc_refresh"):

            # ✅ KEEP LOGIN ONLY
            keep_keys = ["logged_in", "username"]

            # ✅ CLEAR EVERYTHING ELSE
            for key in list(st.session_state.keys()):
                if key not in keep_keys:
                    del st.session_state[key]

            st.rerun()

    st.subheader("Start CDC Reconciliation")

    pdf = st.file_uploader("Upload Consolidated Trades PDF", type=["pdf"], key="cdc")

    if pdf and st.button("Extract & Sort CDC"):
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            tmp.write(pdf.read())
            path = tmp.name

        try:
            tables = camelot.read_pdf(path, pages="all", flavor="stream", edge_tol=500)

            if tables.n == 0:
                st.error("No tables found in PDF.")
                return

            df = pd.concat([t.df for t in tables], ignore_index=True)

            # ===============================
            # STEP 1: REMOVE FIRST 3 ROWS
            # ===============================
            df = df.iloc[3:].reset_index(drop=True)

            # ===============================
            # STEP 2: SET HEADER
            # ===============================
            df.columns = df.iloc[0].astype(str).str.strip()
            df = df.iloc[1:].reset_index(drop=True)

            # Normalize "None"
            df = df.replace("None", "")

            # ===============================
            # STEP 3: SHIFT ZSE LEVY
            # ===============================
            for col in df.columns:
                if "zse levy" in col.lower():
                    df[col] = df[col].astype(str).str.replace(",", "").str.strip()
                    df[col] = pd.to_numeric(df[col], errors="coerce")
                    df[col] = df[col].shift(-1).fillna(0)
                    break

            # ===============================
            # STEP 4: CLEAN FIRST COLUMN
            # ===============================
            first_col = df.columns[0]
            df[first_col] = df[first_col].astype(str).str.strip()

            df = df[
                (df[first_col].notna()) &
                (df[first_col] != "") &
                (~df[first_col].str.lower().isin([
                    "deal number", "broker name", "sell deals"
                ]))
            ]

            # REMOVE rows with "broker code" ANYWHERE

            df = df[
                ~df.apply(
                    lambda row: row.astype(str).str.lower().str.contains("broker code").any(),
                    axis=1
                )
            ]

            # ===============================
            # STEP 5: DROP COLUMNS
            # ===============================
            drop_cols = [
                col for col in df.columns
                if col.lower().strip() in ["deal number", "t+2"]
            ]
            df.drop(columns=drop_cols, inplace=True, errors="ignore")

            # ===============================
            # STEP 6: CLEAN TRADE DATE
            # ===============================
            trade_date_col = safe_find_col(df, ["trade date"])
            if trade_date_col:
                df[trade_date_col] = df[trade_date_col].astype(str).str.strip()
                df = df[
                    (df[trade_date_col] != "") &
                    (df[trade_date_col].str.lower() != "none")
                ]
            else:
                st.warning("Trade Date column not found")

            # ===============================
            # STEP 7: CLEAN QUANTITY
            # ===============================
            qty_col = safe_find_col(df, ["quantity"])
            if qty_col:
                df[qty_col] = pd.to_numeric(
                    df[qty_col].astype(str).str.replace(",", "").str.strip(),
                    errors="coerce"
                ).fillna(0)

                df = df[df[qty_col] > 0]
            else:
                st.warning("Quantity column not found")

            # ===============================
            # STEP 8: NUMERIC CLEANING
            # ===============================
            for col in df.columns:
                if col.lower() in [
                    "trade date", "settlement date",
                    "investor name", "investor code",
                    "counter", "buy/sell"
                ]:
                    continue

                df[col] = pd.to_numeric(
                    df[col].astype(str).str.replace(",", "").str.strip(),
                    errors="coerce"
                )

            df = df.reset_index(drop=True)

            # ===============================
            # STEP 9: BUY / SELL
            # ===============================
            stamp_col = safe_find_col(df, ["stamp duty"])
            qty_col = safe_find_col(df, ["quantity"])

            if not stamp_col or not qty_col:
                st.error("Required columns (Stamp Duty / Quantity) missing.")
                return

            qty_idx = df.columns.get_loc(qty_col)

            bs_col = "Buy/Sell"
            df.insert(qty_idx + 1, bs_col, "")

            df.loc[df[stamp_col] > 0, bs_col] = "BUY"
            df.loc[df[stamp_col] == 0, bs_col] = "SELL"

            buy_df = df[df[stamp_col] > 0].copy()
            sell_df = df[df[stamp_col] == 0].copy()

            # ===============================
            # TOTAL ROW FUNCTION
            # ===============================
            def make_total_row(data, label):
                row = {}
                for c in data.columns:
                    if pd.api.types.is_numeric_dtype(data[c]):
                        row[c] = data[c].sum()
                    else:
                        row[c] = ""
                row[data.columns[0]] = label
                return pd.DataFrame([row])

            # ===============================
            # FINAL CDC TABLE
            # ===============================
            cdc_df = pd.concat([
                pd.concat([buy_df, make_total_row(buy_df, "BUY TOTAL")]),
                pd.concat([sell_df, make_total_row(sell_df, "SELL TOTAL")])
            ], ignore_index=True)

            # ===============================
            # STORE IN SESSION
            # ===============================
            st.session_state.cdc_df = cdc_df
            st.session_state.cdc_sorted = True

            st.success("CDC Processing Complete ✅")
            st.rerun()

        except Exception as e:
            st.error(f"Error extracting tables: {e}")

        finally:
            if os.path.exists(path):
                os.remove(path)

    # ===============================
    # DISPLAY
    # ===============================
    # ---------- Step 2: Show CDC table ----------
    if not st.session_state.get("cdc_sorted", False):
        return

    st.dataframe(st.session_state.cdc_df, use_container_width=True)
    st.divider()

    # ---------- Step 3: Upload Sharestock ----------
    st.subheader("Upload Sharestock File")
    sh_file = st.file_uploader("Upload Sharestock Excel File", type=["xlsx", "xls"], key="sh")

    if sh_file:
        raw = pd.read_excel(sh_file, header=None, engine="openpyxl")

        header_idx = None
        for i, v in raw.iloc[:, 0].items():
            if isinstance(v, str) and v.strip().upper() == "CLIENT":
                header_idx = i
                break

        if header_idx is None:
            st.error("Could not find a 'CLIENT' header row.")
            return

        headers = raw.loc[header_idx]
        clean_headers, seen = [], {}
        for i, h in enumerate(headers):
            name = str(h).strip() if pd.notna(h) and str(h).strip() else f"UNNAMED_{i}"
            if name in seen:
                seen[name] += 1
                name = f"{name}_{seen[name]}"
            else:
                seen[name] = 0
            clean_headers.append(name)

        sh_df = raw.loc[header_idx + 1:].copy()
        sh_df.columns = clean_headers
        sh_df = sh_df.loc[:, ~sh_df.columns.str.startswith("UNNAMED")]
        sh_df = sh_df.dropna(how="all").reset_index(drop=True)

        qty_col = safe_find_col(sh_df, ["qty"])
        if qty_col is None:
            st.error("Qty column not found in Sharestock file.")
            return

        qty_idx = sh_df.columns.get_loc(qty_col)
        for col in sh_df.columns[qty_idx:]:
            sh_df[col] = ensure_numeric(
                sh_df[col].astype(str).str.replace(",", "", regex=False)
            )

        st.session_state.sh_df = sh_df
        st.subheader("Sharestock Data")
        st.dataframe(sh_df, use_container_width=True)

    # ---------- Step 4: Match button ----------
    if st.session_state.get("sh_df") is not None and not st.session_state.get("cdc_matched", False):
        if st.button("Match CDC to Sharestock"):
            _run_cdc_match()

    # ---------- Step 5: Show reconciled tables (persisted) ----------
    if st.session_state.get("cdc_matched", False):
        st.success("CDC ↔ Sharestock matching, totals, and variance complete.")
        st.subheader("Sharestock – Purchases (with CDC + Variance)")
        st.dataframe(st.session_state.purchase_df, use_container_width=True)
        st.subheader("Sharestock – Sales (with CDC + Variance)")
        st.dataframe(st.session_state.sale_df, use_container_width=True)

        if not st.session_state.get("show_final_summary", False):
            if st.button("Final Settlement Summary"):
                st.session_state.show_final_summary = True
                st.rerun()

    # ---------- Step 6: Final settlement summary (persisted) ----------
    if st.session_state.get("show_final_summary", False):
        _show_final_summary()


# =====================================================
# MATCH LOGIC  (called once, results saved to state)
# =====================================================
def take_before_comma(val):
    if pd.isna(val):
        return 0.0
    val_str = str(val).strip()
    if "," in val_str:
        val_str = val_str.split(",")[0]
    return pd.to_numeric(val_str.replace(",", ""), errors="coerce")


def _run_cdc_match():
    cdc = st.session_state.cdc_df.copy()
    sh  = st.session_state.sh_df.copy()

    # ----------------------------
    # FIND COLUMNS
    # ----------------------------
    cdc_type   = safe_find_col(cdc, ["buy/sell", "type"])
    cdc_qty    = safe_find_col(cdc, ["quantity", "qty"])
    sh_type    = safe_find_col(sh,  ["buy/sell", "type"])
    sh_qty     = safe_find_col(sh,  ["quantity", "qty"])
    cdc_sec    = safe_find_col(cdc, ["counter"])
    sh_sec     = safe_find_col(sh,  ["security"])
    cdc_inv    = safe_find_col(cdc, ["investor name", "client"])
    sh_client  = safe_find_col(sh,  ["client"])
    cdc_deal   = get_col_keyword(cdc, ["deal value"])
    sh_deal    = get_col_keyword(sh,  ["gross proceeds"])
    cdc_date   = get_col_keyword(cdc, ["trade date", "settlement date"])
    sh_date    = get_col_keyword(sh,  ["date"])

    if None in (cdc_type, cdc_qty, sh_type, sh_qty, cdc_sec, sh_sec, cdc_inv, sh_client):
        st.error("Missing required columns for matching.")
        return

    # ----------------------------
    # NORMALISE NUMERIC FIELDS
    # ----------------------------
    cdc[cdc_qty] = pd.to_numeric(cdc[cdc_qty], errors="coerce").fillna(0)
    sh[sh_qty]   = pd.to_numeric(sh[sh_qty],   errors="coerce").fillna(0)

    if cdc_deal:
        cdc[cdc_deal] = pd.to_numeric(cdc[cdc_deal], errors="coerce").fillna(0)
    if sh_deal:
        sh[sh_deal]   = pd.to_numeric(sh[sh_deal],   errors="coerce").fillna(0)

    # ----------------------------
    # NORMALISE TYPE
    # CDC uses BUY/SELL → map to PURCHASE/SALE to align with Sharestock.
    # Both sides use the same _MATCH_ key so matching logic is identical
    # for purchases and sales — no separate code paths.
    # ----------------------------
    cdc[cdc_type] = cdc[cdc_type].astype(str).str.upper()
    sh[sh_type]   = sh[sh_type].astype(str).str.upper().replace({
        "BUY":  "PURCHASE",
        "SELL": "SALE",
    })
    # Map CDC BUY → PURCHASE and SELL → SALE (same ladder for both sides)
    cdc["_MATCH_"] = cdc[cdc_type].map({"BUY": "PURCHASE", "SELL": "SALE"})

    # ----------------------------
    # NORMALISE SECURITY TICKER
    # ----------------------------
    KNOWN_MAP = {
        "DELTA":                 "DLTA",
        "TANGANDA":              "TANG",
        "FIRST MUTUAL HOLDINGS": "FML",
        "FIRST MUTUAL":          "FML",
        "CFI":                   "CFI",
        "ZIMRE":                 "ZIMR",
        "MEIKLES":               "MEIK",
        "PROPLASTICS":           "PROL",
        "FBC":                   "FBC",
        "SEED":                  "SEED",
        "TSL":                   "TSL",
        "HIPPO":                 "HIPO",
        "REVITUS":               "REV",
        "DAIRIBORD":               "DZL",
        "ZSETF":                 "ZSEH",
        "ZSE":                   "ZSEH",
        "TIGER":                 "TIG",
        "TURNALL":               "TURN",

        "NATIONAL":              "NMB",
        "AXIA":                  "AXIA",
        "BINDURA":               "BIND",
        "CAFCA":                 "CAFA",
        "COLCOM":                "COLC",
        "DAWN":                  "DAWN",
        "EDGARS":                "EDG",
        "HWANGE":                "HWAN",
        "MEDTECH":               "MED",
        "NMBZ":                  "NMB",
        "OK ZIMBABWE":           "OK",


        "RAINBOW":               "RNB",
        "RIOZIM":                "RIO",
        "SEEDCO Ltd":           "SEED",
        "STARAFRICA":            "STAR",
        "WILLDALE":              "WILL",
        "ZIMPAPERS":             "ZPAP",
        "AFRICAN SUN":           "AFSU",
        "ART":                   "ART",
    }

    sh_tickers = (
        sh[sh_sec]
        .dropna()
        .astype(str)
        .str.upper()
        .str.split(".")
        .str[0]
        .str.strip()
        .unique()
        .tolist()
    )

    def cdc_name_to_ticker(name):
        name_up = str(name).strip().upper()
        for keyword, ticker in sorted(KNOWN_MAP.items(), key=lambda x: len(x[0]), reverse=True):
            if keyword in name_up:
                return ticker
        for ticker in sorted(sh_tickers, key=len, reverse=True):
            if ticker in name_up:
                return ticker
        parts = name_up.split(".")[0].split()
        return parts[0] if parts else ""

    def sh_name_to_ticker(name):
        return str(name).strip().upper().split(".")[0]

    cdc["_SEC_KEY_"] = cdc[cdc_sec].apply(cdc_name_to_ticker)
    sh["_SEC_KEY_"]  = sh[sh_sec].apply(sh_name_to_ticker)

    # Debug expander
    with st.expander("🔍 Security key debug (click to inspect)", expanded=False):
        st.write("**CDC counter → resolved key:**")
        st.dataframe(
            cdc[[cdc_sec, "_SEC_KEY_"]].drop_duplicates().reset_index(drop=True),
            use_container_width=True,
        )
        st.write("**Sharestock security → resolved key:**")
        st.dataframe(
            sh[[sh_sec, "_SEC_KEY_"]].drop_duplicates().reset_index(drop=True),
            use_container_width=True,
        )

    # ----------------------------
    # NORMALISE DATES
    # ----------------------------
    if cdc_date:
        cdc["_DATE_KEY_"] = cdc[cdc_date].apply(norm_date)
    if sh_date:
        sh["_DATE_KEY_"]  = sh[sh_date].apply(norm_date)

    # ----------------------------
    # NORMALISE CLIENT NAME (for Level-3 tiebreaker)
    # ----------------------------
    cdc["_NAME_KEY_"] = cdc[cdc_inv].fillna("").astype(str).str.upper().str.replace(
        r"[^A-Z]", "", regex=True
    )
    sh["_NAME_KEY_"]  = sh[sh_client].fillna("").astype(str).str.upper().str.replace(
        r"[^A-Z]", "", regex=True
    )

    # ----------------------------
    # UNIFIED ROW-BY-ROW MATCH
    #
    # Identical logic runs for BOTH BUY (→ PURCHASE) and SELL (→ SALE) rows.
    # Priority ladder:
    #   L1 – Security + Qty + Type + Date + Deal Value  (all 5 fields)
    #   L2 – Security + Qty + Type + Date               (drop deal value)
    #   L3 – Name tiebreaker when multiple rows survive L1/L2
    #   L4 – First-come, first-served via used_sh_indices (no SH row reuse)
    #
    # The only difference between a BUY and SELL match is the value of
    # _MATCH_ ("PURCHASE" vs "SALE") which filters sh[sh_type] accordingly.
    # There are no separate code branches per side.
    # ----------------------------
    results         = []
    used_sh_indices = set()

    first_col  = cdc.columns[0]
    cdc_clean  = cdc[~cdc[first_col].astype(str).str.contains("TOTAL", na=False)].copy()

    for idx, row in cdc_clean.iterrows():
        match_type = row.get("_MATCH_", "")   # "PURCHASE" or "SALE"
        sec_key    = row.get("_SEC_KEY_", "")
        qty_val    = row[cdc_qty]
        date_key   = row.get("_DATE_KEY_", "") if cdc_date else ""
        name_key   = row.get("_NAME_KEY_", "")

        _raw_deal  = pd.to_numeric(row[cdc_deal], errors="coerce") if cdc_deal else float("nan")
        deal_val   = round(float(_raw_deal), 2) if not pd.isna(_raw_deal) else None
        has_deal   = deal_val is not None

        # ── Sharestock pool for this side (PURCHASE or SALE) ──────────
        sh_pool = sh[sh[sh_type] == match_type]

        # ── L1: Security + Qty + Type + Date + Deal Value ─────────────
        if cdc_date and sh_date and sh_deal and has_deal:
            candidates = sh_pool[
                (sh_pool["_SEC_KEY_"]   == sec_key) &
                (sh_pool[sh_qty]        == qty_val) &
                (sh_pool["_DATE_KEY_"]  == date_key) &
                (sh_pool[sh_deal].round(2) == deal_val)
            ]
        # ── L2: Security + Qty + Type + Date  (drop deal value) ───────
        elif cdc_date and sh_date:
            candidates = sh_pool[
                (sh_pool["_SEC_KEY_"]  == sec_key) &
                (sh_pool[sh_qty]       == qty_val) &
                (sh_pool["_DATE_KEY_"] == date_key)
            ]
        # ── Fallback: no date columns available ────────────────────────
        else:
            candidates = sh_pool[
                (sh_pool["_SEC_KEY_"] == sec_key) &
                (sh_pool[sh_qty]      == qty_val)
            ]
            if sh_deal and has_deal:
                candidates = candidates[candidates[sh_deal].round(2) == deal_val]

        # ── L3: name tiebreaker when multiple candidates survive ───────
        if len(candidates) > 1:
            refined = candidates[candidates["_NAME_KEY_"] == name_key]
            if len(refined) >= 1:
                candidates = refined

        # ── L4: first-come, first-served (no SH row reused) ───────────
        candidates = candidates[~candidates.index.isin(used_sh_indices)]

        if len(candidates) >= 1:
            match_row   = candidates.iloc[0]
            used_sh_indices.add(match_row.name)
            status      = "Matched"
            match_index = match_row.name
        else:
            status      = "Unmatched"
            match_index = None

        results.append({
            "cdc_index":  idx,
            "sh_index":   match_index,
            "status":     status,
            "side":       match_type,   # "PURCHASE" or "SALE" — for diagnostics
        })

    result_df = pd.DataFrame(results)
    st.session_state.match_results = result_df

    # ----------------------------
    # COLUMN MAP  (CDC column keyword → Sharestock column keywords)
    # ----------------------------
    column_map = {
        "quantity":      ["qty"],
        "deal value":    ["gross proceeds"],
        "commission":    ["brokerage"],
        "deal no": ["deal no"],
        "price":["price"],
        "vat":           ["vat"],
        "basic charge":["basic charge"],
        "secz levy":     ["sec levy"],
        "csd levy":      ["csd levy"],
        "zse levy":      ["zse levy"],
        "capital gains": ["cgt"],
        "ipl":           ["investor levy"],
        "stamp duty":    ["stamp duty"],
    }

    # ----------------------------
    # BUILD CDC-ANCHORED OUTPUT ROWS
    #
    # The output schema mirrors Sharestock columns throughout so that
    # downstream summary / export code stays consistent.
    #
    # For EVERY CDC trade (both BUY and SELL):
    #   • Matched  → populate all Sharestock columns from the matched SH row,
    #                then overwrite identity fields (type, security, client,
    #                date, qty) from CDC to guarantee CDC is the authority.
    #   • Unmatched → populate identity columns from CDC; numeric levy
    #                 columns from CDC via column_map; leave SH-only
    #                 columns blank.
    #
    # This is the same routine for purchases AND sales — no branching.
    # ----------------------------
    sh_cols = [c for c in sh.columns if not c.startswith("_")]

    def _sh_col(cdc_key):
        return get_col_keyword(sh, column_map.get(cdc_key, [cdc_key]))

    def _cdc_col(cdc_key):
        return get_col_keyword(cdc, [cdc_key])

    output_rows = []
    for r in results:
        cidx    = r["cdc_index"]
        cdc_row = cdc.loc[cidx]
        sh_row  = sh.loc[r["sh_index"]] if r["sh_index"] is not None else None
        out     = {c: "" for c in sh_cols}

        if sh_row is not None:
            # ── Matched: start from Sharestock row ───────────────
            for c in sh_cols:
                out[c] = sh_row[c] if c in sh_row.index else ""

            # ✅ Only fill missing critical fields from Sharestock
            deal_col = get_col_keyword(sh, ["deal no"])
            price_col = get_col_keyword(sh, ["price"])
            basic_col = get_col_keyword(sh, ["basic charge"])

            for col in [deal_col, price_col, basic_col]:
                if col:
                    val = out.get(col)

                    # If missing → fill from Sharestock row
                    if val is None or val == "" or pd.isna(val):
                        out[col] = sh_row[col]


            # ── Then enforce CDC as authority for identity columns ─────
            # Type
            if sh_type in out:
                out[sh_type] = r["side"]                     # "PURCHASE" or "SALE"
            # Security / Counter
            if sh_sec and cdc_sec:
                out[sh_sec] = cdc_row[cdc_sec]
            # Client / Investor name
            if sh_client and cdc_inv:
                out[sh_client] = cdc_row[cdc_inv]
            # Date
            if sh_date and cdc_date:
                out[sh_date] = cdc_row[cdc_date]
            # Quantity
            if sh_qty and cdc_qty:
                out[sh_qty] = cdc_row[cdc_qty]

        else:
            # ── Unmatched: populate from CDC only ─────────────────────
            # Identity fields
            if sh_type in out:
                out[sh_type]   = r["side"]
            if sh_sec and cdc_sec:
                out[sh_sec]    = cdc_row[cdc_sec]
            if sh_client and cdc_inv:
                out[sh_client] = cdc_row[cdc_inv]
            if sh_date and cdc_date:
                out[sh_date]   = cdc_row[cdc_date]
            if sh_qty and cdc_qty:
                out[sh_qty]    = cdc_row[cdc_qty]

            # Numeric levy / value fields via column_map
            for cdc_key, sh_keys in column_map.items():
                cdc_c = _cdc_col(cdc_key)
                sh_c  = get_col_keyword(sh, sh_keys)
                if cdc_c and sh_c and cdc_c in cdc_row.index:
                    out[sh_c] = cdc_row[cdc_c]

        output_rows.append(out)

    combined = pd.DataFrame(output_rows, columns=sh_cols)

    # Make sure numeric columns are numeric
    non_numeric_cols = {
        sh_type, sh_sec, sh_client, sh_date or "__",
        "Deal No", "deal no",
    }
    for c in combined.columns:
        if c not in non_numeric_cols:
            combined[c] = pd.to_numeric(combined[c], errors="coerce")
    st.write("Sharestock columns:", sh_cols)
    # ----------------------------
    # SPLIT INTO PURCHASES AND SALES
    # (same split logic for both sides — no special-casing)
    # ----------------------------
    combined_buy  = combined[combined[sh_type] == "PURCHASE"].copy()
    combined_sell = combined[combined[sh_type] == "SALE"].copy()

    # ----------------------------
    # TOTALS helper (identical for both sides)
    # ----------------------------
    def add_total(df, label):
        total = {
            col: (df[col].sum() if pd.api.types.is_numeric_dtype(df[col]) else "")
            for col in df.columns
        }
        total[sh_type] = label
        return pd.concat([df, pd.DataFrame([total])], ignore_index=True)

    purchase_df = add_total(combined_buy,  "PURCHASE TOTAL")
    sale_df     = add_total(combined_sell, "SALE TOTAL")

    # ----------------------------
    # CDC TOTAL ROW builder (identical for both sides)
    # ----------------------------
    def build_cdc_total_row(source_df, cdc_subset, label):
        row = {col: "" for col in source_df.columns}
        row[sh_type] = label
        for cdc_key, sh_keys in column_map.items():
            cdc_col = _cdc_col(cdc_key)
            sh_col  = get_col_keyword(source_df, sh_keys)
            if cdc_col and sh_col and cdc_col in cdc_subset.columns:
                val = pd.to_numeric(cdc_subset[cdc_col], errors="coerce").fillna(0)
                row[sh_col] = round(float(val.sum()), 2)
        return pd.DataFrame([row])

    def insert_after(df, after_label, new_row):
        idx_list = df.index[df[sh_type] == after_label].tolist()
        if idx_list:
            idx = idx_list[0] + 1
            return pd.concat([df.iloc[:idx], new_row, df.iloc[idx:]], ignore_index=True)
        return pd.concat([df, new_row], ignore_index=True)

    # All CDC rows per side (exclude TOTAL rows — CDC is the reference)
    cdc_buys  = cdc_clean[cdc_clean["_MATCH_"] == "PURCHASE"]
    cdc_sells = cdc_clean[cdc_clean["_MATCH_"] == "SALE"]

    if not cdc_buys.empty:
        purchase_df = insert_after(
            purchase_df, "PURCHASE TOTAL",
            build_cdc_total_row(purchase_df, cdc_buys, "CDC PURCHASE TOTAL")
        )

    if not cdc_sells.empty:
        sale_df = insert_after(
            sale_df, "SALE TOTAL",
            build_cdc_total_row(sale_df, cdc_sells, "CDC SALE TOTAL")
        )

    # ----------------------------
    # VARIANCE  =  Sharestock TOTAL − CDC TOTAL
    # Same calculation for both sides.
    # ----------------------------
    def add_variance(df, total_label, cdc_label):
        t_rows = df[df[sh_type] == total_label]
        c_rows = df[df[sh_type] == cdc_label]
        if t_rows.empty or c_rows.empty:
            return df
        t = t_rows.iloc[0]
        c = c_rows.iloc[0]
        row = {}
        for col in df.columns:
            if col == sh_type:
                row[col] = "VARIANCE"
            elif pd.api.types.is_numeric_dtype(df[col]):
                tv = pd.to_numeric(t[col], errors="coerce")
                cv = pd.to_numeric(c[col], errors="coerce")
                row[col] = (0 if pd.isna(tv) else tv) - (0 if pd.isna(cv) else cv)
            else:
                row[col] = ""
        return insert_after(df, cdc_label, pd.DataFrame([row]))

    purchase_df = add_variance(purchase_df, "PURCHASE TOTAL", "CDC PURCHASE TOTAL")
    sale_df     = add_variance(sale_df,     "SALE TOTAL",     "CDC SALE TOTAL")

    # ----------------------------
    # SAVE TO SESSION
    # ----------------------------
    st.session_state.purchase_df = purchase_df
    st.session_state.sale_df     = sale_df
    st.session_state.cdc_matched = True

    matched_count   = sum(1 for r in results if r["status"] == "Matched")
    unmatched_count = sum(1 for r in results if r["status"] == "Unmatched")
    buy_matched     = sum(1 for r in results if r["status"] == "Matched"   and r["side"] == "PURCHASE")
    sell_matched    = sum(1 for r in results if r["status"] == "Matched"   and r["side"] == "SALE")
    buy_unmatched   = sum(1 for r in results if r["status"] == "Unmatched" and r["side"] == "PURCHASE")
    sell_unmatched  = sum(1 for r in results if r["status"] == "Unmatched" and r["side"] == "SALE")

    st.info(
        f"Matching complete: **{matched_count} matched** "
        f"(Purchases: {buy_matched} | Sales: {sell_matched}), "
        f"**{unmatched_count} unmatched** "
        f"(Purchases: {buy_unmatched} | Sales: {sell_unmatched})."
    )

    st.rerun()


# =====================================================
# FINAL SETTLEMENT SUMMARY
# =====================================================
def _show_final_summary():
    st.divider()
    st.subheader("Final Settlement Summary")

    csd_reallocation = st.number_input(
        "CSD Reallocation", min_value=0.0, value=0.0, step=0.01, format="%.2f"
    )

    bank_amount = st.number_input(
        "Bank Statement Amount", min_value=0.0, value=0.0, step=0.01, format="%.2f"
    )

    levy_kw = ["broker", "stamp", "zse", "cgt", "sec", "investor", "vat"]

    p = st.session_state.get("purchase_df")
    s = st.session_state.get("sale_df")

    p_type = safe_find_col(p, ["type"]) if p is not None else None
    s_type = safe_find_col(s, ["type"]) if s is not None else None

    total_purchases = extract_levies(p, p_type, "PURCHASE TOTAL", levy_kw) if p_type else 0.0
    total_sales     = extract_levies(s, s_type, "SALE TOTAL",     levy_kw) if s_type else 0.0

    zse_levy = (
        (extract_levy(p, p_type, "PURCHASE TOTAL", "zse") if p_type else 0.0) +
        (extract_levy(s, s_type, "SALE TOTAL",     "zse") if s_type else 0.0)
    )
    ipl_levy = (
        (extract_levy(p, p_type, "PURCHASE TOTAL", "investor") if p_type else 0.0) +
        (extract_levy(s, s_type, "SALE TOTAL",     "investor") if s_type else 0.0)
    )
    sec_levy = (
        (extract_levy(p, p_type, "PURCHASE TOTAL", "sec levy") if p_type else 0.0) +
        (extract_levy(s, s_type, "SALE TOTAL",     "sec levy") if s_type else 0.0)
    )

    total_p_and_s    = round(total_purchases + total_sales, 2)
    total_to_receive = round(total_p_and_s - ipl_levy - sec_levy, 2)
    balance_from_bank = round(total_to_receive - bank_amount, 2)

    post_settlement_total = 0.0
    cdc_df = st.session_state.get("cdc_df")
    if cdc_df is not None:
        fc        = cdc_df.columns[0]
        total_col = next((c for c in cdc_df.columns if "total" in str(c).lower()), None)
        buy_row   = cdc_df[cdc_df[fc].astype(str).str.contains("BUY TOTAL",  case=False, na=False)]
        sell_row  = cdc_df[cdc_df[fc].astype(str).str.contains("SELL TOTAL", case=False, na=False)]
        if total_col and not buy_row.empty and not sell_row.empty:
            b  = pd.to_numeric(buy_row.iloc[0][total_col],  errors="coerce")
            sv = pd.to_numeric(sell_row.iloc[0][total_col], errors="coerce")
            post_settlement_total = round(
                (0 if pd.isna(b)  else float(b)) +
                (0 if pd.isna(sv) else float(sv)) -
                zse_levy - ipl_levy - sec_levy,
                2,
            )

    cgt_variance = 0.0
    if s is not None and s_type:
        vr = s[s[s_type] == "VARIANCE"]
        if not vr.empty:
            cgt_col = next((c for c in vr.columns if "cgt" in c.lower()), None)
            if cgt_col:
                val = pd.to_numeric(vr.iloc[0][cgt_col], errors="coerce")
                cgt_variance = 0.0 if pd.isna(val) else round(float(val), 2)

    diff_bank_post = round(balance_from_bank - cgt_variance - csd_reallocation, 2)
    verify_value   = round(
        bank_amount + cgt_variance + csd_reallocation + diff_bank_post - post_settlement_total,
        2,
    )

    summary_df = pd.DataFrame([
        ["Total Purchases",              total_purchases,    "", "", "", "", "", "", ""],
        ["Total Sales",                  total_sales,        "", "", "", "", "", "", ""],
        ["Total Purchases and Sales",    total_p_and_s,      "", "", "", "", "", "", ""],
        ["",                             "",                 "", "", "", "", "", "", ""],
        ["IPL Levy Remitted Directly",   ipl_levy,           "", "", "", "", "", "", ""],
        ["SEC Levy",                     sec_levy,           "", "", "", "", "", "", ""],
        [
            "", "", "",
            "Bank Statement Amount",
            "Balance from Amount to be Received and Bank Amount",
            "Capital Gains",
            "CSD Reallocation ",
            "Balance Bank Charges ",
            "Post Settlement Total from Report",
            "Verify",
            "", "",
        ],
        [
            "Total Amount to be Received",
            total_to_receive,
            "",
            bank_amount,
            balance_from_bank,
            cgt_variance,
            csd_reallocation,
            diff_bank_post,
            post_settlement_total,
            verify_value,
            "",
        ],
        ["", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", ""],
        ["Journals to be Posted in Sharestock", "", "", ""],
        ["LEVIES REMITTED DIRECTLY BY DEPOSITORY", "", "", ""],
        ["Narration", "", "DR", "CR"],
        ["CDC Bank charges", "BANK CHARGES", diff_bank_post, "", "", "", "", "", ""],
        ["", "FBC BANK", "", diff_bank_post],
        ["", "", "", ""],
        ["Total", "", diff_bank_post, diff_bank_post],
        ["", "", "", ""],
        ["", "", "", ""],
        ["", "", "", ""],
        ["Prepared by:....................................................................................................", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", ""],
        ["Reviewed by:....................................................................................................", "", "", "", "", "", "", "", ""],
    ])

    st.dataframe(summary_df, use_container_width=True)
    _export_excel(summary_df, post_settlement_total)


# =====================================================
# EXCEL EXPORT
# =====================================================
def _export_excel(summary_df, post_settlement_total):
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        st.session_state.cdc_df.to_excel(writer, sheet_name="CDC Raw", index=False)
        st.session_state.sh_df.to_excel(writer, sheet_name="Sharestock Raw", index=False)
        st.session_state.purchase_df.to_excel(writer, sheet_name="Purchases", index=False)

        SALES_SHEET = "Sales & Summary"
        st.session_state.sale_df.to_excel(writer, sheet_name=SALES_SHEET, index=False, startrow=0)
        summary_df.to_excel(
            writer, sheet_name=SALES_SHEET,
            index=False, header=False,
            startrow=len(st.session_state.sale_df) + 3,
        )

        bold  = Font(bold=True)
        thick = Side(style="medium")
        thin  = Side(style="thin")

        # ---- Sales & Summary sheet ----
        ws = writer.sheets[SALES_SHEET]

        for cell in ws[1]:
            cell.font = bold

        for r in range(2, ws.max_row + 1):
            val = ws.cell(row=r, column=1).value
            if val is None or str(val).strip() == "":
                for rr in range(r, ws.max_row + 1):
                    cell = ws.cell(row=rr, column=1)
                    cell.value = cell.value or ""
                    cell.font  = bold
                break

        rows_hit = []
        for r in range(ws.max_row, 1, -1):
            if ws.cell(row=r, column=17).value not in (None, "", " "):
                rows_hit.append(r)
            if len(rows_hit) == 3:
                break
        for r in rows_hit:
            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).font = bold

        for r in range(1, ws.max_row + 1):
            if str(ws.cell(row=r, column=1).value or "").strip().lower() == "total sales":
                ws.cell(row=r, column=1).border = Border(bottom=thick)
                ws.cell(row=r, column=2).border = Border(bottom=thick)
                break

        for r in range(1, ws.max_row + 1):
            if str(ws.cell(row=r, column=1).value or "").strip().lower() == "total amount to be received":
                ws.cell(row=r, column=1).border = Border(top=thick, bottom=Side(style="double"))
                ws.cell(row=r, column=2).border = Border(top=thick, bottom=Side(style="double"))
                break

        thin = Side(style="thin")
        bold = Font(bold=True)

        journal_start = None
        for r in range(1, ws.max_row + 1):
            val = str(ws.cell(row=r, column=1).value or "").strip().lower()
            if "journals to be posted in" in val:
                journal_start = r
                break

        if journal_start:
            header_row = journal_start + 2
            data_start = header_row + 1
            total_row  = data_start + 3

            for r in range(header_row, total_row + 1):
                for c in range(1, 5):
                    cell = ws.cell(row=r, column=c)
                    cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

            for c in range(1, 5):
                ws.cell(row=header_row, column=c).font = bold
            for c in range(1, 5):
                ws.cell(row=total_row,  column=c).font = bold

            for r in range(header_row, total_row + 1):
                ws.cell(row=r, column=3).alignment = Alignment(horizontal="right")
                ws.cell(row=r, column=4).alignment = Alignment(horizontal="right")

        bank_headers = {
            "bank statement amount",
            "balance from amount to be received and bank amount",
            "capital gains",
            "post settlement total from report",
            "balance bank charges",
            "csd reallocation",
            "verify",
        }
        header_row, found_cols = None, []
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                if str(ws.cell(row=r, column=c).value or "").strip().lower() in bank_headers:
                    header_row = r
                    break
            if header_row:
                break
        if header_row:
            for c in range(1, ws.max_column + 1):
                if str(ws.cell(row=header_row, column=c).value or "").strip().lower() in bank_headers:
                    found_cols.append(c)
            if found_cols:
                sc, ec = min(found_cols), max(found_cols)
                vr = header_row + 1
                for row_r in (header_row, vr):
                    for c in range(sc, ec + 1):
                        ws.cell(row=row_r, column=c).border = Border(
                            top=thick    if row_r == header_row else None,
                            bottom=thick if row_r == vr         else None,
                            left=thick,
                            right=thick,
                        )
                for c in range(sc, ec + 1):
                    ws.cell(row=vr, column=c).border = Border(
                        top=thin, bottom=thin, left=thin, right=thin
                    )

        for r in range(1, ws.max_row + 1):
            if str(ws.cell(row=r, column=1).value or "").strip().lower() == "sec levy":
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r + 1, column=c).alignment = Alignment(
                        wrap_text=True, vertical="bottom"
                    )
                break

        # ---- Purchases sheet ----
        wp = writer.sheets["Purchases"]
        for cell in wp[1]:
            cell.font = bold
        for r in range(max(1, wp.max_row - 2), wp.max_row + 1):
            for c in range(1, wp.max_column + 1):
                cell = wp.cell(row=r, column=c)
                if "sharestock" not in str(cell.value or "").lower():
                    cell.font = bold

        # ---- Global: Times New Roman 12, number format, auto widths ----
        num_fmt = '#,##0.00;-#,##0.00;-'
        for wsh in writer.book.worksheets:
            for row in wsh.iter_rows():
                for cell in row:
                    existing  = cell.font or Font()
                    cell.font = Font(
                        name="Times New Roman", size=12,
                        bold=existing.bold, italic=existing.italic,
                    )
                    if isinstance(cell.value, str):
                        try:
                            cell.value = float(cell.value.replace(",", ""))
                        except Exception:
                            pass
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = num_fmt
                        cell.alignment     = Alignment(horizontal="right", vertical="center")

            for col_idx in range(1, wsh.max_column + 1):
                col_letter = get_column_letter(col_idx)
                max_len    = max(
                    (len(str(cell.value)) for cell in wsh[col_letter] if cell.value is not None),
                    default=0,
                )
                wsh.column_dimensions[col_letter].width = min(28, max(8, max_len + 2))

        writer.book.active = 0

    output.seek(0)
    file_name = f"CDC_RECEIPTING_{post_settlement_total:.2f}.xlsx"
    file_path = os.path.join(HISTORY_FOLDER, file_name)

    if st.download_button(
        label="📥 Save & Download CDC Report",
        data=output,
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ):
        with open(file_path, "wb") as f:
            f.write(output.getvalue())

        history = load_history()
        history.append({
            "id":    str(uuid.uuid4()),
            "file":  file_name,
            "user":  st.session_state.get("username", ""),
            "date":  datetime.now().strftime("%Y-%m-%d %H:%M"),
            "type":  "CDC Receipting",
            "total": float(post_settlement_total),
        })
        save_history(history)
        st.success("CDC report saved to history.")
